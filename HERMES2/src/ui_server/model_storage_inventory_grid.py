#!/usr/bin/env python
# -*- coding: utf-8 -*-
###################################################################################
# Copyright   2015, Pittsburgh Supercomputing Center (PSC).  All Rights Reserved. #
# =============================================================================== #
#                                                                                 #
# Permission to use, copy, and modify this software and its documentation without # 
# fee for personal use within your organization is hereby granted, provided that  #
# the above copyright notice is preserved in all copies and that the copyright    # 
# and this permission notice appear in supporting documentation.  All other       #
# restrictions and obligations are defined in the GNU Affero General Public       #
# License v3 (AGPL-3.0) located at http://www.gnu.org/licenses/agpl-3.0.html  A   #
# copy of the license is also provided in the top level of the source directory,  #
# in the file LICENSE.txt.                                                        #
#                                                                                 #
###################################################################################
_hermes_svn_id_="$Id$"

import sys,os,os.path,time,json,math,types
import bottle
import ipath
import site_info
from serverconfig import rootPath
from upload import uploadAndStore, makeClientFileInfo
import typehelper
from typeholdermodel import allTypesModelName
import shadow_network_db_api
import privs
import session_support_wrapper as session_support
import shadow_network as shd
import util
from HermesServiceException import HermesServiceException
import crumbtracks
from modelhooks import addCrumb

from ui_utils import _logMessage, _logStacktrace, _safeGetReqParam, _getOrThrowError, _smartStrip, b64E, b64D

## If executing under mod_wsgi, we need to add the path to the source
## directory of this script.
sI = site_info.SiteInfo()

inlizer=session_support.inlizer
_=session_support.translateString

@bottle.route('/model-edit-store-inventory-tabular')
def storeInvTabPage(db,uiSession):
    crumbTrack = addCrumb(uiSession,_("Edit Population Counts (Tabular)"))
    try:
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId)        
        
        #popTypes = [(x,m.types[x].DisplayName) for x in m.types.keys() if isinstance(m.types[x],shd.ShdPeopleType)]
        
        return bottle.template("model_storeinv_edit.tpl",
                               {"breadcrumbPairs":crumbTrack,
                                'modelId':modelId
                               })
    except bottle.HTTPResponse:
        raise 

@bottle.route('/test-store-device-list')
def testStoreDevList(db,uiSession):
    try:
        crumbTrack = addCrumb(uiSession,_("Test Store Inventory Grid Page"))
        modelId =  _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        storeId = _getOrThrowError(bottle.request.params,'storeId',isInt=True)
        
        return bottle.template("test_store_inventory_grid.tpl",
                               {'breadcrumbPairs':crumbTrack,
                                'modelId':modelId,'storeId':storeId})
    
    except bottle.HTTPResponse:
        raise 
       
@bottle.route('/json/get-storage-devices-for-model-for-storeId')
def getDeviceInventoryListForModelForStoreIdjQGridJSON(db,uiSession):
    import json
    try:
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        storeId = _getOrThrowError(bottle.request.params,'storeId',isInt=True)
        m = shadow_network_db_api.ShdNetworkDB(db,modelId)  
        #print m.stores.keys[100000]
        rowList = []
        
        if storeId not in m.stores.keys():
            return {'success':False, 
                    'type':'error', 
                    'msg':_('StoreId {0} not in model {1} set of stores'.format(storeId,modelId))}
        
        for dev,count in m.stores[storeId].countAllInventory().items():
            if isinstance(m.types[dev],shd.ShdStorageType):
                rowList.append({'device':m.types[dev].getDisplayName(),
                                'count':count})
                 
        results = {
                   'success':True,
                   'total':1,
                   'page':1,
                   'records':len(rowList),
                   'rows':rowList
                   }
    
        return results
    
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        result = {'success':False, 'type':'error', 'msg':str(e)}
        return result  
          
@bottle.get('/json/manage-store-inventory-grid')
def jsonManagePopStoreGrid(db,uiSession):
    import json
    try: 
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId)    
    
        result = createStoreInvGridJSonFromModel(m)   
        
        result['success'] = True
    
        return result

    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        result = {'success':False, 'type':'error','msg':str(e)}
        return result             
        
### somehow need to make this use display names
def createStoreInvGridJSonFromModel(m):
    popTypes = [x for x in m.types.keys() if isinstance(m.types[x],shd.ShdPeopleType)]
    rowList = []
    sortRowList = sorted([x for x,y in m.stores.items() if not y.isAttached()])
    for storeId in sortRowList:
        store = m.stores[storeId]
        
        row = {}
        row['idcode'] = storeId
        row['name'] = store.NAME
        row['level'] = store.CATEGORY 
        row['inventory'] = storeId
        ## Note the inventory and edit buttons will be rendered automatically by the widget
        rowList.append(row)
    
    results = {
               'total':1,
               'page':1,
               'records':len(rowList),
               'rows':rowList
               }
    
    return results
        

@bottle.route('/edit/verify-edit-population-storegrid',method='post')
def jsonStoreGridPopulationVerifyAndCommit(db,uiSession):
    try:
        if bottle.request.params['oper'] == 'edit':
            modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
            idcode  = _getOrThrowError(bottle.request.params,'idcode',isInt=True)
            
            uiSession.getPrivs().mayModifyModelId(db, modelId)
            m = shadow_network_db_api.ShdNetworkDB(db,modelId)
            
            popTypes = [x for x in m.types.keys() if isinstance(m.types[x],shd.ShdPeopleType)]
            
            newPopValues = {x:_getOrThrowError(bottle.request.params,x,isInt=True) for x in popTypes}
            
            store = m.stores[idcode]
            for k,v in newPopValues.items():
                store.updateDemand(k,v)
                
            db.commit()
            
            return {'success':True}
        elif bottle.request.params['oper']=='add':
            raise bottle.BottleException(_('jsonStoreGridPopulationVerifyAndCommit: unsupported operation add'))
        elif bottle.request.params['oper']=='del':
            raise bottle.BottleException(_('jsonStoreGridPopulationVerifyAndCommit: unsupported operation del'))
    
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        print str(e)
        result = {'success':False, 'type':'error','msg':str(e)}
        return result  
    
@bottle.route('/edit/verify-edit-population-storegrid-from-json',method='post')
def jsonStoreGridGeoCoordVerifyAndCommitFromJson(db, uiSession):
    import json
    try:    
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        jsonStrToUpdate = _getOrThrowError(bottle.request.params,'jsonStr',isInt=False)
        jsonToUpdate = json.loads(jsonStrToUpdate)
         
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db,modelId)
        #print "JSON!!!!! = {0}".format(jsonToUpdate)
        for updateItem in jsonToUpdate:
            idcode = updateItem[0]
            storeToUpdate = m.stores[idcode]
            for uI in updateItem[1:]:
                storeToUpdate.updateDemand(uI[0],uI[1])
                #print "US = {0}".format(uI)
            #storeToUpdate.Latitude = updateItem[1]
            #storeToUpdate.Longitude = updateItem[2]
         
        db.commit()
             
        return {'success':True}
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        print str(e)
        result = {'success':False, 'type':'error','msg':str(e)}
        return result  

def createPopulationTemplateSpreadsheet(m):
    from openpyxl import Workbook
    from results_excel_report import XLCell,plainStyle
    try:
        #modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        #uiSession.getPrivs().mayModifyModelId(db, modelId)
        #m = shadow_network_db_api.ShdNetworkDB(db, modelId)  
    
        jsonList = createPopStoreGridJSonFromModel(m)
        popTypes = [x for x in m.types.keys() if isinstance(m.types[x],shd.ShdPeopleType)]
        wb = Workbook()
        sheet = wb.create_sheet(0)
        sheet.title = _("Model Population Estimates")
        
        ss = XLCell(sheet)
        ss.postNext(_('HERMES ID'),plainStyle)
        ss.postNext(_("Location Name"),plainStyle)
        ss.postNext(_("Supply Chain Level"),plainStyle)
        ss.postNext(_("Attached Demand?"),plainStyle)
        
        
        for pT in popTypes:
            ss.postNext(pT,plainStyle)
        
        #ss.postNext(_("Latitude"),plainStyle)
        #ss.postNext(_("Longitude"),plainStyle)
        
        ss.nextRow()
        
        for row in jsonList['rows']:
            #print row['idcode']
            ss.postNext(row['idcode'],plainStyle)
            ss.postNext(row['name'],plainStyle)
            ss.postNext(row['level'],plainStyle)
            ss.postNext(row['attached'],plainStyle)
            for pT in popTypes:
                ss.postNext(row[pT],plainStyle)
            #ss.postNext(row['latitude'],plainStyle)
            #ss.postNext(row['longitude'],plainStyle)
            ss.nextRow()
        
        return wb
    
    except Exception,e:
        raise

@bottle.route('/downloadTemplatePopulationXLS')
def downloadTemplatePopulationXLS(db,uiSession):        
    try:
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId)    
        
        wb = createPopulationTemplateSpreadsheet(m)
        xlsName = "model_{0}_poptemplat.xlsx".format(modelId)
        with uiSession.getLockedState() as state:
            (fileKey, fullFileName) = \
                state.fs().makeNewFileInfo(shortName=xlsName,
                                           fileType='application/vnd.ms-excel',
                                           deleteIfPresent=True)
        wb.save(fullFileName)
        resp = bottle.static_file(xlsName, os.path.dirname(fullFileName),
                                  mimetype='application/vnd.ms-excel', download=xlsName)
        resp.set_cookie("fileDownload","true",path='/')
        return resp
        #bottle.response.set_cookie("fileDownload","true",path='/')
        #print bottle.response
        #return bottle.static_file(xlsName, os.path.dirname(fullFileName),
        #                          mimetype='application/vnd.ms-excel', download=xlsName)
    except Exception as e:
        print 'Exception: %s'%e
        raise



def validatePopulationSpreadsheet(fileName,m):
    from collections import Counter
    from openpyxl import load_workbook
    returnDict = {'success':True,'badNames':[],'dups':[],'updates':[],'message':''}
    try:
        wb = load_workbook(filename=fileName)
        ws = wb["Model Population Estimates"]
        popTypes = [x for x in m.types.keys() if isinstance(m.types[x],shd.ShdPeopleType)]
    except Exception,e:
        returnDict['success'] = False
        returnDict['message'] = 'Cannot load notebook {0}'.format(fileName)
        return returnDict
    
    # Check Headers
    if not (ws['A1'].value == "HERMES ID" and \
            ws['B1'].value == "Location Name" and \
            ws['C1'].value == "Supply Chain Level" and \
            ws['D1'].value == "Attached Demand?"):
            #ws["D1"].value == "Latitude" and \
            #ws['E1'].value == "Longitude"):
        returnDict['success'] = False
        returnDict['message'] = "Invalid Population Spreadsheet: Headers of spreadsheet do not conform"
        return returnDict
    
    # To Do next.
    col = 5
    validateCells = []
    for i in range(col,col+(len(popTypes))):
        validateCells.append(ws.cell(row=1,column=i).value)
    print "!!!!! validate Cells = {0}".format(validateCells)
    for vC in validateCells:
        if vC not in popTypes:
            returnDict['success'] = False
            returnDict['message'] = "Invalid Population Spreadsheet: Population Type {0} not in the current model".format(vC)
            return returnDict
    
    badNames = []
    dupNames = []
    updateList = []
    modelNameTuples = [(x.idcode,x.NAME, x.CATEGORY) for x in m.stores.values()]
        
    spreadNameDict = {(x[0].value,x[1].value,x[2].value):[(popTypes[y],x[y+4].value) for y in range(0,len(popTypes))]\
                       for x in ws.iter_rows(row_offset=1) if x[0].value is not None}
    ## Find Bad Names
    
    for name in spreadNameDict.keys():
        if name not in modelNameTuples:
            badNames.append(name)
                    

    countsList = Counter(spreadNameDict.keys())
    for elem,count in countsList.items():
        if count > 1:
            dupNames.append(elem)
    
    if len(badNames) > 0:
        returnDict['badNames'] = badNames
        returnDict['success'] = True
    
    if len(dupNames) > 0:
        returnDict['dups'] = dupNames
        returnDict['success'] = True
    
    
    for name,pops in spreadNameDict.items():
        if name not in badNames and name not in dupNames:
            updateList.append(tuple([name[0]]+pops))
    
    returnDict['updates'] = updateList
    return returnDict
        
@bottle.post('/upload-populationspreadsheet')
def uploadGeoCoordSpreadsheet(db,uiSession):
    fileKey = None
    try:
        
        info = uploadAndStore(bottle.request, uiSession)
        modelId = info['modelId']
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId) 
        clientData = makeClientFileInfo(info)
        clientData['code']= 0
        clientData['message'] = ''
        clientData['files'] = [{
                               'name':info['shortName'],
                               'size':os.path.getsize(info['serverSideName'])
                               }]
        
        validDict = validatePopulationSpreadsheet(info['serverSideName'], m)
        print validDict
        clientData['validResult'] = validDict
        
        return json.dumps(clientData)

    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        _logMessage("\nException: %s"%e)
        _logStacktrace()
        # We could return text which would go into a browser window set as the 'target'
        # for this upload, but there doesn't seem to be much point.
        if fileKey:
            try:
                with uiSession.getLockedState() as state:
                    try:
                        os.remove(info['serverSideName'])
                    except:
                        pass
                    state.fs().removeFileInfo(fileKey)
            except:
                pass
        db.rollback()
        mdata = { 'code': 1, 'message': str(e) }            
        return json.dumps(mdata)   
    
    
                   