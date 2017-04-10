#!/usr/bin/env python
# -*- coding: utf-8 -*-
###################################################################################
# Copyright   2015, Pittsburgh Supercomputing Center (PSC).  All Rights Reserved. #
# =============================================================================== #
#                                                                                 #
# Permission to use, copy, and modify this software and its documentation without # 
# fee for personal use within your organization is hereby granted, provided that  #
# the above copyright notice is preserved in all copies and that the copyright    # 
# and this permission notice appear in supporting documentation.  All other       #
# restrictions and obligations are defined in the GNU Affero General Public       #
# License v3 (AGPL-3.0) located at http://www.gnu.org/licenses/agpl-3.0.html  A   #
# copy of the license is also provided in the top level of the source directory,  #
# in the file LICENSE.txt.                                                        #
#                                                                                 #
###################################################################################
_hermes_svn_id_="$Id$"

import sys,os,os.path,time,json,math,types
import bottle
from sqlalchemy.exc import SQLAlchemyError
import ipath
import site_info
from serverconfig import rootPath
from preordertree import PreOrderTree
from upload import uploadAndStore, makeClientFileInfo
import htmlgenerator
import typehelper
from typeholdermodel import allTypesModelName
import shadow_network_db_api
import privs
import session_support_wrapper as session_support
import shadow_network as shd
import util
from HermesServiceException import HermesServiceException
from gridtools import orderAndChopPage
from model_edit_hooks import updateDBRouteType
import crumbtracks
import runhooks
from modelhooks import addCrumb

from ui_utils import _logMessage, _logStacktrace, _safeGetReqParam, _getOrThrowError, _smartStrip, b64E, b64D

## If executing under mod_wsgi, we need to add the path to the source
## directory of this script.
sI = site_info.SiteInfo()

inlizer=session_support.inlizer
_=session_support.translateString

@bottle.route('/model-edit-geocoords-tabular')
def geoCoordTes(db,uiSession):
    crumbTrack = addCrumb(uiSession, _("Edit Geographic Coordinates (Tabular)"))
    try:
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId)        
        
        return bottle.template("model_geocoords_edit.tpl",
                               {"breadcrumbPairs":crumbTrack,
                               'modelId':modelId})
    except bottle.HTTPResponse:
        raise 
    
@bottle.route('/json/manage-geocoord-storegrid')
def jsonManageGeoCoordStorGrid(db,uiSession):
    try:
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId)    
    
        result = createGeoCoordStorGridJSonFromModel(m)
        
        result['success'] = True
        return result
    
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        result = {'success':False, 'type':'error','msg':str(e)}
        return result             

def createGeoCoordStorGridJSonFromModel(m):
    rowList = []
    sortRowList = sorted([x for x,y in m.stores.items()])
    for storeId in sortRowList:
        store = m.stores[storeId]
        if not store.isAttached() and not store.isSurrogate() and not store.isOutreachClinic():
            row = {}
            row['idcode'] = storeId
            row['name'] = store.NAME
            row['level'] = store.CATEGORY
            if store.hasGIS():
                row['latitude'] = store.Latitude
                row['longitude'] = store.Longitude
                row['oldlatitude'] = store.Latitude
                row['oldlongitude'] = store.Longitude
            else:
                row['latitude'] = ""
                row['longitude'] = ""
                row['oldlatitude'] = ""
                row['oldlongitude'] = ""
            
            #row['modelId'] = modelId
            rowList.append(row)
        
    results = {
               'total':1,
               'page':1,
               'records':len(rowList),
               'rows':rowList
               }
    
    return results
    
@bottle.route('/edit/verify-edit-geocoord-storegrid',method='post')
def jsonStoreGridGeoCoordVerifyAndCommit(db, uiSession):
    try:
        if bottle.request.params['oper'] == 'edit':
            modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
            idcode = _getOrThrowError(bottle.request.params,'idcode',isInt=True)
            latitude = _getOrThrowError(bottle.request.params,'latitude',isInt=False)
            longitude = _getOrThrowError(bottle.request.params,'longitude',isInt=False)
            
            uiSession.getPrivs().mayModifyModelId(db, modelId)
            m = shadow_network_db_api.ShdNetworkDB(db,modelId)
            
            store = m.stores[idcode]
            store.Latitude = latitude
            store.Longitude = longitude
            
            db.commit()
            
            return {'success':True}
        elif bottle.request.params['oper']=='add':
            raise bottle.BottleException(_('unsupported operation'))
        elif bottle.request.params['oper']=='del':
            raise bottle.BottleException(_('unsupported operation'))
    
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        print str(e)
        result = {'success':False, 'type':'error','msg':str(e)}
        return result  

@bottle.route('/edit/verify-edit-geocoord-storegrid-from-json',method='post')
def jsonStoreGridGeoCoordVerifyAndCommitFromJson(db, uiSession):
    import json
    try:    
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        jsonStrToUpdate = _getOrThrowError(bottle.request.params,'jsonStr',isInt=False)
        jsonToUpdate = json.loads(jsonStrToUpdate)
        
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db,modelId)
        for updateItem in jsonToUpdate:
            idcode = updateItem[0]
            storeToUpdate = m.stores[idcode]
            storeToUpdate.Latitude = updateItem[1]
            storeToUpdate.Longitude = updateItem[2]
        
        db.commit()
            
        return {'success':True}
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        print str(e)
        result = {'success':False, 'type':'error','msg':str(e)}
        return result  

def createGeoCoordTemplateSpreadsheet(m):
    from openpyxl import Workbook
    from results_excel_report import XLCell,plainStyle
    try:
        #modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        #uiSession.getPrivs().mayModifyModelId(db, modelId)
        #m = shadow_network_db_api.ShdNetworkDB(db, modelId)  
    
        jsonList = createGeoCoordStorGridJSonFromModel(m)
        
        wb = Workbook()
        sheet = wb.create_sheet(0)
        sheet.title = _("Model GeoCoordinates")
        
        ss = XLCell(sheet)
        ss.postNext(_('HERMES ID'),plainStyle)
        ss.postNext(_("Location Name"),plainStyle)
        ss.postNext(_("Supply Chain Level"),plainStyle)
        ss.postNext(_("Latitude"),plainStyle)
        ss.postNext(_("Longitude"),plainStyle)
        
        ss.nextRow()
        
        for row in jsonList['rows']:
            #print row['idcode']
            ss.postNext(row['idcode'],plainStyle)
            ss.postNext(row['name'],plainStyle)
            ss.postNext(row['level'],plainStyle)
            ss.postNext(row['latitude'],plainStyle)
            ss.postNext(row['longitude'],plainStyle)
            ss.nextRow()
        
        return wb
    
    except Exception,e:
        raise


@bottle.route('/downloadTemplateGeoCoordXLS')
def downloadTemplateGeoCoordXLS(db,uiSession):        
    try:
        modelId = _getOrThrowError(bottle.request.params,'modelId',isInt=True)
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId)    
        
        wb = createGeoCoordTemplateSpreadsheet(m)
        xlsName = "model_{0}_geotemplat.xlsx".format(modelId)
        with uiSession.getLockedState() as state:
            (fileKey, fullFileName) = \
                state.fs().makeNewFileInfo(shortName=xlsName,
                                           fileType='application/vnd.ms-excel',
                                           deleteIfPresent=True)
        wb.save(fullFileName)
        resp = bottle.static_file(xlsName, os.path.dirname(fullFileName),
                                  mimetype='application/vnd.ms-excel', download=xlsName)
        resp.set_cookie("fileDownload","true",path='/')
        return resp
        #bottle.response.set_cookie("fileDownload","true",path='/')
        #print bottle.response
        #return bottle.static_file(xlsName, os.path.dirname(fullFileName),
        #                          mimetype='application/vnd.ms-excel', download=xlsName)
    except Exception as e:
        print 'Exception: %s'%e
        raise
    
def validateGeoCoordSpreadsheet(fileName,m):
    from collections import Counter
    from openpyxl import load_workbook
    returnDict = {'success':True,'badNames':[],'dups':[],'updates':[],'message':''}
    try:
        wb = load_workbook(filename=fileName)
        ws = wb["Model GeoCoordinates"]
    
    except Exception,e:
        returnDict['success'] = False
        returnDict['message'] = 'Cannot load notebook {0}'.format(fileName)
        return returnDict
    
    # Check Headers
    if not (ws['A1'].value == "HERMES ID" and \
            ws['B1'].value == "Location Name" and \
            ws['C1'].value == "Supply Chain Level" and \
            ws["D1"].value == "Latitude" and \
            ws['E1'].value == "Longitude"):
        returnDict['success'] = False
        returnDict['message'] = "Headers of spreadsheet do not conform"
        return returnDict
    
    badNames = []
    dupNames = []
    updateList = []
    modelNameTuples = [(x.idcode,x.NAME, x.CATEGORY) for x in m.stores.values()]
    spreadNameDict = {(x[0].value,x[1].value,x[2].value):(x[3].value,x[4].value)\
                       for x in ws.iter_rows(row_offset=1) if x[0].value is not None}
    ## Find Bad Names
    
    for name in spreadNameDict.keys():
        if name not in modelNameTuples:
            badNames.append(name)
                    

    countsList = Counter(spreadNameDict.keys())
    for elem,count in countsList.items():
        if count > 1:
            dupNames.append(elem)
    
    if len(badNames) > 0:
        returnDict['badNames'] = badNames
        returnDict['success'] = True
    
    if len(dupNames) > 0:
        returnDict['dups'] = dupNames
        returnDict['success'] = True
    
    for name,latlons in spreadNameDict.items():
        if name not in badNames and name not in dupNames:
            updateList.append((name[0],latlons[0],latlons[1]))
    
    returnDict['updates'] = updateList
    return returnDict
        
@bottle.post('/upload-geocoordspreadsheet')
def uploadGeoCoordSpreadsheet(db,uiSession):
    fileKey = None
    try:
        
        info = uploadAndStore(bottle.request, uiSession)
        modelId = info['modelId']
        uiSession.getPrivs().mayModifyModelId(db, modelId)
        m = shadow_network_db_api.ShdNetworkDB(db, modelId) 
        clientData = makeClientFileInfo(info)
        clientData['code']= 0
        clientData['message'] = ''
        clientData['files'] = [{
                               'name':info['shortName'],
                               'size':os.path.getsize(info['serverSideName'])
                               }]
        
        validDict = validateGeoCoordSpreadsheet(info['serverSideName'], m)
        print validDict
        clientData['validResult'] = validDict
        
        return json.dumps(clientData)

    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        _logMessage("\nException: %s"%e)
        _logStacktrace()
        # We could return text which would go into a browser window set as the 'target'
        # for this upload, but there doesn't seem to be much point.
        if fileKey:
            try:
                with uiSession.getLockedState() as state:
                    try:
                        os.remove(info['serverSideName'])
                    except:
                        pass
                    state.fs().removeFileInfo(fileKey)
            except:
                pass
        db.rollback()
        mdata = { 'code': 1, 'message': str(e) }            
        return json.dumps(mdata)   