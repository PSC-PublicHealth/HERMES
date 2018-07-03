#!/usr/bin/env python


from openpyxl.styles import alignment

####################################
# Notes:
# -A session could be hijacked just by grabbing the SessionID;
#  should use an encrypted cookie to prevent this.
####################################
_hermes_svn_id_ = "$Id: resultshooks.py 1927 2014-09-25 21:45:00Z welling $"

import sys, os, os.path, time, json, math, types
import bottle
import ipath
import shadow_network_db_api
import shadow_network as shd
import session_support_wrapper as session_support
from HermesServiceException import HermesServiceException
import privs
import htmlgenerator
import typehelper
from costmodel import getCostModelSummary
from ui_utils import _logMessage, _logStacktrace, _safeGetReqParam, _getOrThrowError
if 0:
    from xlwt.Workbook import *
    from xlwt.Style import *
    import xlwt
    ezxf = xlwt.easyxf

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font,Color,fills
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension

import time
from reporthooks import createModelSummaryWSCall, createModelSummarySerialized
import traceback
from resultshooks import costsSummaryKey

from collections import defaultdict

inlizer = session_support.inlizer
_ = session_support.translateString


class XLCell:
    def __init__(self, worksheet, row=1, column=1):
        self.worksheet = worksheet
        self._row = row
        self._column = column
        self._maxRow = row
        self._maxCol = column

        self.RowHeightCache = {}
        self.ColumnWidthCache = {}


    def ws(self):
        return self.worksheet

    def r(self, row = None):
        if row is not None:
            self._row = row
            if self._row > self._maxRow:
                self._maxRow = self._row
        return self._row

    def c(self, column = None):
        if column is not None:
            self._column = column
            if self._column > self._maxCol:
                self._maxCol = self._column
        return self._column

    def maxRow(self, maxRow=None):
        if maxRow is not None:
            self._maxRow = maxRow
        return self._maxRow

    def maxColumn(self, maxColumn=None):
        if maxColumn is not None:
            self_maxCol = maxColumn
        return self._maxCol

    def cell(self, column = None, row = None):
        if column is not None:
            self.c(column)
        if row is not None:
            self.r(row)
        return self.worksheet.cell(row=self._row, column=self._column)

    def right(self, cols=1):
        self.c(self.c() + cols)
        return self.cell()

    def down(self, rows=1):
        self.r(self.r() + rows)
        return self.cell()

    def left(self, cols=1):
        return self.right(-cols)

    def up(self, rows=1):
        return self.down(-rows)

    def set(self, value=None, style=None):
        c = self.cell()
        if value is not None:
            c.value = value
        if style is not None:
            c.style = style
        return c

    def next(self, value = None, style = None):
        c = self.right()
        if value is not None:
            c.value = value
        if style is not None:
            c.style = style
        return c
    
    def postNext(self, value = None, style = None):
        c = self.cell()
        if value is not None:
            c.value = value
        if style is not None:
            c.style = style
        self.right()
        return c
    
    def nextRow(self):
        self._column = 1
        self.down()
        return self.cell()

    def mergeCells(self, columns=1, rows=1):
        self.ws().merge_cells(start_row=self.r(),
                              start_column=self.c(),
                              end_row=self.r() + rows - 1,
                              end_column=self.c() + columns - 1)

    def columnWidth(self, width, column=None):
        if column is None:
            column = self.c()
        if width not in self.ColumnWidthCache:
            self.ColumnWidthCache[width] = ColumnDimension(worksheet=self.worksheet, width=width)
        colStr = get_column_letter(column)
        self.ws().column_dimensions[colStr] = self.ColumnWidthCache[width]

    def rowHeight(self, height, row=None):
        if row is None:
            row = self.r()
        if height not in self.RowHeightCache:
            self.RowHeightCache[height] = RowDimension(worksheet=self.worksheet, height=height)
        self.ws().row_dimensions[row] = self.RowHeightCache[height]

medBorder = Border(bottom=Side(style='medium', color='FF000000'), 
                   left=Side(style='medium', color='FF000000'), 
                   right=Side(style='medium', color='FF000000'), 
                   top=Side(style='medium', color='FF000000'))


modelHeadingStyle = Style(font=Font(bold=True, color="FF000000", sz=18.0),
                          fill=PatternFill(patternType=fills.FILL_SOLID,
                                           fgColor=Color('FFFFFFFF')), #fgColor=Color('FF911313')),
                          alignment=Alignment(horizontal='left',
                                              vertical='center',
                                              wrap_text=True),
                          border=Border(bottom=Side(color="FF000000", border_style=None)))

modelHeading2Style = Style(font=Font(bold=True, color="FF000000"),
                          fill=PatternFill(patternType=fills.FILL_SOLID,fgColor=Color('FFFFFFFF')),
                          alignment=Alignment(horizontal='left',
                                              vertical='center',
                                            wrap_text=True),
                          border=Border(bottom=Side(color="FF000000", border_style=None)))

divider1Style = Style(font=Font(bold=True, color="FFFFFFFF"),
                      fill=PatternFill(patternType=fills.FILL_SOLID,
                                       fgColor=Color('FF0000FF')),
                      alignment=Alignment(horizontal='left',
                                          vertical='bottom',
                                          wrap_text=True),
                      border=Border(bottom=Side(color="FF000000", 
                                                border_style=None)))

divider2Style = Style(font=Font(bold=True, color="FFFFFFFF"),
                      fill=PatternFill(patternType=fills.FILL_SOLID,
                                       fgColor=Color('FF3709EF')),
                      alignment=Alignment(horizontal='left',
                                          vertical='bottom',
                                          wrap_text=True),
                      border=Border(bottom=Side(color="FF000000", 
                                                border_style=None)))

labelStyle = Style(font=Font(bold=True, color="FF000000"),
                          fill=PatternFill(patternType=fills.FILL_SOLID,
                                           fgColor=Color('FFA0A8EC')),
                          alignment=Alignment(horizontal='left',
                                              vertical='center',
                                              wrap_text=True),
                          border=Border(bottom=Side(color="FF000000", 
                                                    border_style=None)))

labelStyleCenter = Style(font=Font(bold=True, color="FF000000"),
                         fill=PatternFill(patternType=fills.FILL_SOLID,
                                          fgColor=Color('FFA0A8EC')),
                         alignment=Alignment(horizontal='center',
                                             vertical='center',
                                             wrap_text=True),
                         #border=None
                         )

levelStyle = Style(font=Font(bold=True),
                   # fill=PatternFill(fill_type=None,start_color='dark_red',
                   #  end_color='dark_red'),
                   alignment=Alignment(horizontal='left',
                                       vertical='center'))   

levelTotalStyle = levelStyle
totalStyle = Style(font=Font(bold=True),
                   # fill=PatternFill(fill_type=None,start_color='dark_red',
                   #  end_color='dark_red'),
                   alignment=Alignment(horizontal='right',
                                       vertical='center'))   



plainStyle = Style(font=Font(bold=False),
                   # fill=PatternFill(fill_type=None,start_color='dark_red',end_color='dark_red'),
                   alignment=Alignment(horizontal='right',
                                       vertical='center'))   
plainCenteredStyle = Style(font=Font(bold=False),
                   # fill=PatternFill(fill_type=None,start_color='dark_red',end_color='dark_red'),
                   alignment=Alignment(horizontal='center',
                                       vertical='center'))
plainLabelStyle = Style(font=Font(bold=False),
                   # fill=PatternFill(fill_type=None,start_color='dark_red',end_color='dark_red'),
                        alignment=Alignment(horizontal='left',
                                            vertical='center'))
plainTextStyle = plainLabelStyle

def preprocessCostRows(hr):
    csrs = hr.costSummaryRecs

    topEntries = {}
    allEntries = {}
    routeEntries = {}
    storeEntries = {}

    currency = None
    year = None


    for csr in csrs:
        if currency is None:
            currency = csr.Currency
        else:
            assert currency == csr.Currency, "cost summary currencies aren't consistent"
        if year is None:
            year = csr.BaseYear
        else:
            assert year == csr.BaseYear, "cost summary base years aren't consistent"
        rl = csr.ReportingLevel
        rb = csr.ReportingBranch
        if rl == '-top-':
            topEntries[rb] = csr
        elif rl == 'all':
            allEntries[rb] = csr
        elif rb.startswith('loc: '):
            (x, paren, loc) = rb.rpartition('(')
            assert paren == '(', "invalid cost summary rec: %s"%csr
            (loc, paren, x) = loc.partition(')')
            assert paren == ')', "invalid cost summary rec: %s"%csr
            loc = long(loc)
            storeEntries[loc] = csr
        elif rb.startswith('rt: '):
            rt = rb[4:]
            routeEntries[rt] = csr
        else:
            assert False, "unknown cost summary rec type: %s"%csr
        
    return {'top' : topEntries,
            'all' : allEntries,
            'route' : routeEntries,
            'store' : storeEntries,
            'currency' : currency,
            'year' : year }




@bottle.route('/json/create-xls-summary-openpyxl')
def createExcelSummaryOpenPyXl(db, uiSession):
    try:
        import json
        #createModelSummaryWSCall(db, uiSession)
        fname = _safeGetReqParam(bottle.request.params, 'filename', isInt=False)
        modelId = _safeGetReqParam(bottle.request.params, 'modelId', isInt=False)
        resultsId = _safeGetReqParam(bottle.request.params, 'resultsId', isInt=False)
        uiSession.getPrivs().mayReadModelId(db,modelId)
        m = shadow_network_db_api.ShdNetworkDB(db,modelId)
        #modelJSON = json.loads(createModelSummarySerialized(m))
        #hr = m.getResultById(resultsId)
        hr = db.query(shd.HermesResults).filter(shd.HermesResults.resultsId==resultsId).one()
        
        wb = createExcelSummaryOpenPyXLForResult(uiSession,m,hr)
        
        with uiSession.getLockedState() as state:
            (fileKey, fullFileName) = \
                state.fs().makeNewFileInfo(shortName="%s.xlsx" % fname,
                                           fileType='application/vnd.ms-excel',
                                           deleteIfPresent=True)
            
        wb.save(fullFileName)
        return {'success':True,'xlsfilename':fullFileName}
    
    except Exception as e:
        print 'Exception: %s'%e
        traceback.print_exc()

        return {'success':False, 'msg':str(e)}    

def createExcelSummaryOpenPyXLForResult(uiSession, m,hr):
    
    import json
    
    modelJSON = json.loads(createModelSummarySerialized(m))
    levelSummary = summarizeByLevel(hr)
    costs = preprocessCostRows(hr)
    srs = hr.storesRpts

    
    wb = Workbook()
    summary_sheet = wb.create_sheet(0)
    summary_sheet.title = "Model Summary"

    ss = XLCell(summary_sheet)

    ss.mergeCells(5)
    ss.rowHeight(22)
    print modelJSON['name']
    ss.postNext(modelJSON['name'], modelHeadingStyle)

    ss.nextRow()
    ss.mergeCells(5)
    ss.postNext(_("Overall System Statistics"), modelHeading2Style)
    
    ss.nextRow()
    ss.mergeCells(5)
    ss.postNext(_("Facility Statistics"), divider1Style)
    
    ss.nextRow()

    #levels = modelJSON['orderedLevelList']
    #levels = m.getLevelList()
    levels = m.getParameterValue('levellist')
    levelCount = modelJSON['fixedLocationCount']
    vaccLevelCount = modelJSON['vaccinationLocationCount']
    
    ss.rowHeight(45)
    headers = ["Level", "Total", "Vaccinating", "Average Peak Storage Utilization", "Total Volume Delivered (L)"]
    for header in headers:
        ss.postNext(header, labelStyle)

    ss.nextRow()
    
    for level in levels:
        if level in levelCount.keys():
            ss.postNext(level, levelStyle)
            ss.postNext(levelCount[level], plainStyle)
            if level in vaccLevelCount.keys():
                ss.postNext(vaccLevelCount[level], plainStyle)
            else:
                ss.postNext(0, plainStyle)
            ss.postNext(round(levelSummary['averagePeakUtilization'][level], 3), plainStyle)
            ss.postNext(round(levelSummary['deliveryVol'][level], 3), plainStyle)
            #ss.postNext(levelSummary['coolerVol'][level], plainStyle)
            #ss.postNext(levelSummary['usedCoolerVol'][level], plainStyle)
            ss.nextRow()

    ss.postNext(_("All Levels"), levelTotalStyle)
    ss.postNext(levelCount['Total'], totalStyle)
    ss.postNext(vaccLevelCount['Total'], totalStyle)
    ss.postNext(round(levelSummary['averagePeakUtilization']['Total'],3), totalStyle)
    ss.postNext(round(levelSummary['deliveryVol']['Total'], 3), totalStyle)
    #ss.postNext(levelSummary['coolerVol']['Total'], totalStyle)
    #ss.postNext(levelSummary['usedCoolerVol']['Total'], totalStyle)
    vCols = [(_("Vaccine"), "Name", 'label'),
             (_("Availability"), "SupplyRatio", '%'),
             (_("Doses Needed"), "Applied", None),
             (_("Doses Received"), "Treated", None),
             (_("Open Vial Waste"), "OpenVialWasteFrac", '%'),
             (_("Vials Opened"), "VialsUsed", None),
             (_("Vials Procured"), "VialsCreated", None),
             (_("Vials Spoiled"), "VialsExpired", None),
             #(_("Vials Broken"), "VialsBroken"),
             (_("% Stored 2 to 8 C"), "coolerStorageFrac", '%'),
             (_("% Stored Below 2 C"), "freezerStorageFrac", '%')]
    ss.nextRow()
    ss.nextRow()
    ss.mergeCells(len(vCols))
    ss.set(_("Vaccine Statistics"), divider1Style)
    ss.nextRow()
    ss.rowHeight(30)
    for header, attr, fmt in vCols:
        ss.postNext(header, labelStyle)
    ss.nextRow()
    totals = {}    
    for rec in hr.summaryRecs.values():
        if not isinstance(rec, shd.ShdVaccineSummary):
            continue
        for header, attr, fmt in vCols:
            if attr == "Name":
                val = m.types[rec.Name].getDisplayName()
            else:
                val = getattr(rec, attr)
                if attr not in totals.keys():
                    totals[attr] = 0.0
                totals[attr] += val
                    
            if fmt == '%':
                val = round(100.0 * val, 3)
                
            style = plainStyle
            if fmt == 'label':
                style = levelStyle
            
            ss.postNext(val, style)
        ss.nextRow()

    ss.postNext("Totals",levelTotalStyle)
    ss.postNext(round(100.0 * (totals["Treated"]/totals["Applied"]),3),totalStyle)
    ss.postNext(totals["Applied"],totalStyle)
    ss.postNext(totals["Treated"],totalStyle)
    ss.postNext("-",totalStyle)
    ss.postNext(totals["VialsUsed"],totalStyle)
    ss.postNext(totals['VialsCreated'],totalStyle)
    ss.postNext(totals['VialsExpired'],totalStyle)
    ss.postNext("-",totalStyle)
    ss.postNext("-",totalStyle)
    ss.nextRow()
    


    popLevelCount = modelJSON['populationLevelCount']
    numColumns = len(popLevelCount)
    
    ss.nextRow()
    ss.mergeCells(1 + 4*numColumns)
    ss.postNext(_("Population Statistics"), divider1Style)

    ss.nextRow()

    ## Get an ordered list to make all consistent
    popCats = sorted([x for x in popLevelCount['Total'].keys()])

    levelList = []
    for level in levels:
        if level in popLevelCount.keys():
            levelList.append((level,level))
    levelList.append((_("All Levels"), "Total"))

    statList = ((_("Total"), 'count'),
                (_("Average"), 'ave'),
                (_("Minimum"), 'min'),
                (_("Maximum"), 'max'))

    ss.set(style=labelStyleCenter)
    ss.down()
    ss.set(style=labelStyleCenter)
    ss.up()
    ss.right()
    for title, level in levelList:
        ss.mergeCells(len(statList))
        ss.set(title, labelStyleCenter)
        ss.down()
        for statLabel, stat in statList:
            ss.postNext(statLabel, labelStyle)
        ss.up()
    
    ss.nextRow()
    ss.nextRow()

    for cat in popCats:
        ss.postNext(cat, levelStyle)
        for title, level in levelList:
            print "Level = {0} and count = {1}".format(level,popLevelCount)
            for statLabel, stat in statList:
                if cat in popLevelCount[level].keys():
                    val = round(popLevelCount[level][cat][stat])
                else:
                    val = 0
                ss.postNext(val, plainStyle)
        ss.nextRow()


    ss.nextRow()
    row = ss.r()
    col=1
    ss.mergeCells(10)
    ss.set(_("Transport Route Statistics"), divider2Style)
    ss.nextRow()
    row+=1
    
    ## Get an ordered list to make all consistent
    routeLevelDict = modelJSON['routeLevelCount']
    ss.postNext(style=labelStyleCenter)
    ss.mergeCells(4)
    ss.set(_("Distance (KM)"), labelStyleCenter)
    ss.right(4)
    ss.mergeCells(2)
    ss.set(_("Route Types"), labelStyleCenter)
    ss.right(2)
    ss.mergeCells(3)
    ss.set(_("Vehicles"), labelStyleCenter)
    ss.nextRow()

    for label in (_("Level"), 
                  _("Total"), _("Average"), _("Minimum"), _("Maximum"),
                  _("Type"), _("Count"),
                  _("Type"), _("Count"), _("Average Peak % Capacity Needed")):
        ss.postNext(label, labelStyle)

    ss.nextRow()
    
    truckLevelSummary = summarizeTrucksByLevel(m, hr)

    for level in levels:
        if level not in routeLevelDict.keys():
            continue
        ss.postNext(level, levelStyle)
        for stat in ('total', 'ave', 'min', 'max'):
            ss.postNext(round(routeLevelDict[level]['distance'][stat], 2), plainStyle)
        
        rTypeCounts = routeLevelDict[level]['routeTypeCount']
        tTypeCounts = routeLevelDict[level]['vehicleCount']
        for (routeItem, truckItem) in map(None,
                                          rTypeCounts.items(),
                                          tTypeCounts.items()):
            if routeItem is not None:
                ss.postNext(routeItem[0], plainTextStyle)
                ss.postNext(routeItem[1], plainStyle)
            else:
                ss.right(2)
            if truckItem is not None:
                ss.postNext(m.types[truckItem[0]].getDisplayName(), plainTextStyle)
                ss.postNext(truckItem[1], plainStyle)
                try:
                    fill = truckLevelSummary['fill'][level][truckItem[0]]
                    count = truckLevelSummary['count'][level][truckItem[0]]
                    fillPercent = round(fill * 100.0 / count, 3)
                except:
                    fillPercent = ""
                ss.postNext(fillPercent, plainStyle)
            else:
                ss.right(3)
            ss.left(5)
            ss.down()

        ss.cell(column=1)

    ss.nextRow()
    
    if len(hr.costSummaryRecs) > 0:
        if isinstance(hr.costSummaryRecs[0], shd.ShdMicro1CostSummaryGrp):
            summarizeMicroCostLevels(costs, levels, ss)

        elif isinstance(hr.costSummaryRecs[0], shd.ShdLegacyCostSummary):
            ss.set("legacycostsummary")
        else:
            ss.set("unknown cost summary")
        
        ss.nextRow()



    ss.mergeCells(4)
    ss.set(_("Cost Per Dose and Fully Immunized Child") + " (%s %s)"%(costs['year'], costs['currency']), divider1Style)
    ss.nextRow()
    ss.mergeCells(2)
    ss.postNext(_("Cost Per Dose"), labelStyleCenter)
    ss.next()
    ss.mergeCells(2)
    ss.postNext(_("Cost Per FIC"), labelStyleCenter)
    ss.nextRow()
    ss.rowHeight(30)
    ss.postNext(_("Logistics"), labelStyle)
    ss.postNext(_("with Procurement"), labelStyle)
    ss.postNext(_("Logistics"), labelStyle)
    ss.postNext(_("with Procurement"), labelStyle)
    ss.nextRow()

    keyCosts = costsSummaryKey(m, hr)
    ss.postNext(round(keyCosts['logCostPerDose'], 3), plainStyle)
    ss.postNext(round(keyCosts['costPerDose'], 3), plainStyle)
    ss.postNext(round(keyCosts['logCostPerFIC'], 3), plainStyle)
    ss.postNext(round(keyCosts['costPerFIC'], 3), plainStyle)
    

    ss.columnWidth(20, column=1)
    for i in xrange(2, ss.maxColumn()):
        ss.columnWidth(16, column=i)


    dev_inventory_sheet = wb.create_sheet(1)
    dev_inventory_sheet.title = _("Level-wise Device Inventory")

    dis = XLCell(dev_inventory_sheet)
    dis.columnWidth(20, column=1)
    dis.columnWidth(30, column=2)
    dis.columnWidth(15, column=3)


    dis.mergeCells(3)
    dis.set(_("Device Inventory by Level"), divider1Style)
    dis.nextRow()
    dis.rowHeight(30)
    for label in (_("Level"), _("Device"), _("Number of Devices")):
        dis.postNext(label, labelStyle)
    
    dis.nextRow()
    
    inventoryCount = modelJSON['inventoryLevelCount']
    
    for level in levels:
        if level in inventoryCount.keys():
            dis.set(level, levelStyle)
            for dev,count in inventoryCount[level].items():
                dis.next(m.types[dev].getDisplayName(), plainLabelStyle)
                dis.next(count, plainStyle)
                dis.nextRow()
    
                
                
    
    heir_sheet = wb.create_sheet(2)
    heir_sheet.title = "Store Locations"

    hs = XLCell(heir_sheet)

    heirInfo = modelJSON['heirarchicalStores']
    hs.nextRow()
    hs.nextRow()
    hs.nextRow()

    for placeDict in heirInfo:
        hs.right(placeDict['depth'])
        hs.postNext(placeDict['name'], plainLabelStyle)
        hs.nextRow()

    levCol = hs.maxColumn()

    mcd = MicroCostDisplay('stores')
    

    hs.r(1)
    hs.mergeCells(levCol + 7)
    hs.set(_("Storage Locations"), divider1Style)
    hs.c(levCol+8)
    hs.mergeCells(mcd.columnCount)
    hs.set(_("Costs") + " (%s %s)"%(costs['year'], costs['currency']), divider1Style)
    hs.nextRow()
    hs.rowHeight(30)
    hs.mergeCells(levCol,2)
    hs.set("Name", labelStyleCenter)
    hs.c(levCol+1)
    headers = (_("Level"), _("ID Code"), 
               _("Latitude"), _("Longitude"),
               _("Straight KM Distance"),
               _("Total Volume Delivered"),
               _("Peak Storage Utilization"),
               _("Vaccine Availability"),)

    for header in headers:
        hs.mergeCells(1, 2)
        hs.postNext(header, labelStyleCenter)

    costsCol = hs.c()
    mcd.topHeaders(hs, labelStyleCenter)
    hs.nextRow()
    
    hs.rowHeight(30)
    hs.c(costsCol)
    mcd.detailHeaders(hs, labelStyle)
    hs.nextRow()

    for placeDict in heirInfo:
        code = placeDict['idcode']
        dep = placeDict['depth']
        if dep > 1:
            hs.mergeCells(dep)
        hs.right(dep)
        hs.mergeCells(levCol- dep)
        hs.c(levCol+1)
        hs.postNext(placeDict['level'], plainLabelStyle)
        hs.postNext(code, plainLabelStyle)
        if placeDict['latitude'] != 0.0 and placeDict['longitude'] != 0.0:
            hs.postNext(placeDict['latitude'], plainStyle)
            hs.postNext(placeDict['longitude'], plainStyle)
            hs.postNext(placeDict['distKM'],plainStyle)
        else:
            hs.next()
            hs.next()
            hs.next()

        if code in srs:
            sr = srs[code]
            hs.postNext(round(sr.tot_delivery_vol, 3), plainStyle)
            if 'cooler' in sr.storage and sr.storage['cooler'].vol > 0.0:
                cooler = sr.storage['cooler']
                hs.postNext(round(cooler.fillRatio * 100.0, 3), plainStyle)
            else:
                hs.next()
            patients = 0
            treated = 0
            for vr in sr.vax.values():
                patients += vr.patients
                treated += vr.treated
            if patients == 0:
                hs.next()
            else:
                hs.postNext(round(100.0 * treated / patients, 3), plainStyle)
        else:
            hs.next()
            hs.next()

        if code in costs['store']:
            mcd.detailRow(hs, costs['store'][code], plainStyle)
        

        hs.nextRow()

    for i in xrange(1,levCol):
        hs.columnWidth(6, column=i)
    hs.columnWidth(20, column=levCol)
    hs.columnWidth(20, column=levCol+1)
    for i in xrange (levCol+2, hs.maxColumn()+1):
        hs.columnWidth(16, column=i)


    route_sheet = wb.create_sheet(3)
    route_sheet.title = "Routes"
    rs = XLCell(route_sheet)
    mcd = MicroCostDisplay('routes')

    rs.mergeCells(8) # need to set this
    rs.set(_("Transport Routes"), divider1Style)
    rs.right(8)
    rs.mergeCells(mcd.columnCount)
    rs.set(_("Costs") + " (%s %s)"%(costs['year'], costs['currency']), divider1Style)
    rs.nextRow()

    rs.mergeCells(1, 2)
    rs.postNext(_('Route'), labelStyle)
    rs.mergeCells(2)
    rs.postNext(_('Supplier'), labelStyleCenter)
    rs.next()
    rs.mergeCells(2)
    rs.postNext(_('Recipient'), labelStyleCenter)
    rs.next()
    rs.mergeCells(1, 2)
    rs.postNext(_('Vehicle'), labelStyle)
    rs.mergeCells(1, 2)
    rs.postNext(_('Trips'), labelStyle)
    rs.mergeCells(1, 2)
    rs.postNext(_('Peak % Capacity Needed'), labelStyle)

    costsCol = rs.c()
    mcd.topHeaders(rs, labelStyle)
    rs.nextRow()

    rs.rowHeight(30)
    rs.next()
    rs.postNext(_('Name'), labelStyle)
    rs.postNext(_('Level'), labelStyle)
    rs.postNext(_('Name'), labelStyle)
    rs.postNext(_('Level'), labelStyle)
    rs.c(costsCol)
    mcd.detailHeaders(rs, labelStyle)
    rs.nextRow()

    rRpts = hr.routesRpts

    for routeId, route in m.routes.items():
        if route.Type == "attached":
            continue

        rs.postNext(routeId, levelStyle)
        supplier = route.supplier()
        clients = route.clients()

        rs.postNext(supplier.NAME, plainLabelStyle)
        rs.postNext(supplier.CATEGORY, plainTextStyle)
        rs.postNext(clients[0].NAME, plainLabelStyle)
        rs.postNext(clients[0].CATEGORY, plainTextStyle)
        rs.postNext(m.types[route.TruckType].getDisplayName(), plainTextStyle)
        if routeId in rRpts:
            rpt = rRpts[routeId]
            rs.postNext(rpt.RouteTrips, plainStyle)
            rs.postNext(round(rpt.RouteFill*100.0,3), plainStyle)
        else:
            rs.next()
            rs.next()
        if routeId in costs['route']:
            mcd.detailRow(rs, costs['route'][routeId], plainStyle)
        rs.nextRow()
        for i in xrange(1, len(clients)):
            rs.c(4)
            rs.postNext(clients[i].NAME, plainLabelStyle)
            rs.postNext(clients[i].CATEGORY, plainTextStyle)
            rs.nextRow()

    for i in xrange(1, 7):
        rs.columnWidth(20, column=i)
    for i in xrange(7, rs.maxColumn()):
        rs.columnWidth(16, column=i)
    
    storage_sheet = wb.create_sheet(4)
    storage_sheet.title = "Storage Devices"
    s = XLCell(storage_sheet)
    
    s.mergeCells(9)
    s.set(_("Storage Devices"), divider1Style)
    s.nextRow()

    s.mergeCells(1,2)
    s.postNext(_("Device"), labelStyle)
    s.mergeCells(2)
    s.postNext(_("Net Capacity (L)"), labelStyleCenter)
    s.next()
    s.mergeCells(3)
    s.postNext(_("Price"), labelStyleCenter)
    s.next()
    s.next()
    s.mergeCells(1,2)
    s.postNext(_("Amortization Years"), labelStyle)
    s.mergeCells(2)
    s.postNext(_("Energy"), labelStyleCenter)
    s.nextRow()

    s.rowHeight(30)
    s.next()
    s.postNext(_("2 to 8 C"), labelStyle)
    s.postNext(_("Below 2 C"), labelStyle)
    s.postNext(_("Amount"), labelStyle)
    s.postNext(_("Currency"), labelStyle)
    s.postNext(_("Year"), labelStyle)
    s.next()
    s.postNext(_("Usage"), labelStyle)
    s.postNext(_("Units"), labelStyle)
    s.nextRow()

    for f in m.fridges.values():
        s.postNext(m.types[f.Name].getDisplayName(), plainLabelStyle)
        s.postNext(f.cooler, plainStyle)
        s.postNext(f.freezer, plainStyle)
        s.postNext(f.BaseCost, plainStyle)
        s.postNext(f.BaseCostCurCode, plainStyle)
        s.postNext(f.BaseCostYear, plainStyle)
        s.postNext(f.AmortYears, plainStyle)
        s.postNext(f.PowerRate, plainStyle)
        s.postNext(f.PowerRateUnits, plainStyle)
        s.nextRow()

    s.columnWidth(20, column=1)
    for i in xrange(2, rs.maxColumn()):
        s.columnWidth(12, column=i)


    vehicle_sheet = wb.create_sheet(5)
    vehicle_sheet.title = "Vehicles"
    s = XLCell(vehicle_sheet)

    s.mergeCells(9)
    s.set(_("Vehicles"), divider1Style)
    s.nextRow()


    s.mergeCells(1,2)
    s.postNext(_("Vehicle"), labelStyle)
    s.mergeCells(1,2)
    s.postNext(_("2-8 C Capacity (L)"), labelStyle)
    s.mergeCells(3)
    s.postNext(_("Price"), labelStyleCenter)
    s.next()
    s.next()
    s.mergeCells(1,2)
    s.postNext(_("Amortization km"), labelStyle)
    s.mergeCells(3)
    s.postNext(_("Fuel"), labelStyleCenter)
    s.nextRow()

    s.rowHeight(30)
    s.next()
    s.next()
    s.postNext(_("Amount"), labelStyle)
    s.postNext(_("Currency"), labelStyle)
    s.postNext(_("Year"), labelStyle)
    s.next()
    s.postNext(_("Type"), labelStyle)
    s.postNext(_("Usage"), labelStyle)
    s.postNext(_("Units"), labelStyle)
    s.nextRow()

    for t in m.trucks.values():
        s.postNext(m.types[t.Name].getDisplayName(), plainLabelStyle)
        s.postNext(t.totalCoolVolume(m), plainStyle)
        s.postNext(t.BaseCost, plainStyle)
        s.postNext(t.BaseCostCurCode, plainStyle)
        s.postNext(t.BaseCostYear, plainStyle)
        s.postNext(t.AmortizationKm, plainStyle)
        s.postNext(t.Fuel, plainStyle)
        s.postNext(t.FuelRate, plainStyle)
        s.postNext(t.FuelRateUnits, plainStyle)
        s.nextRow()

    s.columnWidth(20, column=1)
    for i in xrange(2, rs.maxColumn()):
        s.columnWidth(12, column=i)


    vaccine_sheet = wb.create_sheet(6)
    vaccine_sheet.title = "Vaccines"
    s = XLCell(vaccine_sheet)

    s.mergeCells(12)
    s.set(_("Vaccines"), divider1Style)
    s.nextRow()


    s.mergeCells(1,2)
    s.postNext(_("Vaccine"), labelStyle)
    s.mergeCells(1,2)
    s.postNext(_("Presentation"), labelStyle)
    s.mergeCells(1,2)
    s.postNext(_("Doses Per Vial"), labelStyle)
    s.mergeCells(2)
    s.postNext(_("Packaged mL Per Dose"), labelStyleCenter)
    s.next()
    s.mergeCells(4)
    s.postNext(_("Potent Lifetime (Days)"), labelStyleCenter)
    s.right(3)
    s.mergeCells(3)
    s.postNext(_("Price Per Vial"), labelStyleCenter)
    #############################
    #  Doses Per Person Go Here #
    #############################
    s.nextRow()
    s.rowHeight(30)
    s.right(3)
    s.postNext(_("Vaccine"), labelStyle)
    s.postNext(_("Diluent"), labelStyle)
    s.postNext(_("2 to 8 C"), labelStyle)
    s.postNext(_("Below 2 C"), labelStyle)
    s.postNext(_("After Opening"), labelStyle)
    s.postNext(_("Out of Cold Chain"), labelStyle)
    s.postNext(_("Amount"), labelStyle)
    s.postNext(_("Currency"), labelStyle)
    s.postNext(_("Year"), labelStyle)
    s.nextRow()

    for v in m.vaccines.values():
        s.postNext(m.types[v.Name].getDisplayName(), plainLabelStyle)
        s.postNext(v.presentation, plainTextStyle)
        s.postNext(v.dosesPerVial, plainStyle)
        s.postNext(v.volPerDose, plainStyle)
        s.postNext(v.diluentVolPerDose, plainStyle)
        s.postNext(v.coolerLifetime, plainStyle)
        s.postNext(v.freezerLifetime, plainStyle)
        s.postNext(v.openLifetime, plainStyle)
        s.postNext(v.roomtempLifetime, plainStyle)
        s.postNext(v.pricePerVial, plainStyle)
        s.postNext(v.priceUnits, plainStyle)
        s.postNext(v.priceBaseYear, plainStyle)
        s.nextRow()



    s.columnWidth(30, column=1)
    s.columnWidth(20, column=2)
    for i in xrange(3, rs.maxColumn()):
        s.columnWidth(12, column=i)

               

    return wb
    #return {'success':True, 'xlsfilename':fullFileName}
    
#    except Exception as e:
#        print 'Exception: %s'%e
#        traceback.print_exc()

#        return {'success':False, 'msg':str(e)}
    
# @bottle.route('/json/create-xls-summary')
# def createExcelSummary(db, uiSession):
#     try:
#         modelJSON = createModelSummary(db, uiSession)
#     except Exception as e:
#         return {'success':False, 'msg':str(e)}
#     
#     try:
#         fname = _safeGetReqParam(bottle.request.params, 'filename', isInt=False)
#         # ## First create a summary page
#         style = XFStyle()
#         wb = Workbook()
#         ws0 = wb.add_sheet('Summary')
#     
#         row = 0
#         col = 1
#         # Define some formatting styles for the spreadsheet.
#         modelHeading_xf = ezxf('font: bold on, height 240; align: wrap on, horiz centre, vert top; borders: bottom thin')
#         divider_1_xf = ezxf('font: bold on, height 222, colour white; align: horiz left; pattern: pattern solid, fore-color dark_red; borders: bottom thin')
#         divider_2_xf = ezxf('font: bold on, height 222, colour white; align: horiz left; pattern: pattern solid, fore-color light_blue; borders: bottom thin')
#         total_label_xf = ezxf('font: bold on, height 222; align: horiz left; pattern: pattern solid, fore-color grey25', num_format_str='0%')
#         total_percent_xf = ezxf('font: bold on, height 222; align: horiz right; pattern: pattern solid, fore-color grey25', num_format_str='0%')
#         total_xf = ezxf('font: bold on, height 222; align: horiz right; pattern: pattern solid, fore-color grey25', num_format_str='#,##0')
#         plain_xf = ezxf('font: bold off, height 222; align: horiz right', num_format_str="#,##0")
#         plain_label_xf = ezxf('font: bold off, height 222; align: horiz left')
#         plain_percent_xf = ezxf('font: bold off, height 222; align: horiz right', num_format_str='0%')
#         doses_label_xf = ezxf('font: bold on, height 222; align: horiz left; pattern: pattern solid, fore-color olive_ega', num_format_str="#,##0")
#         doses_xf = ezxf('font: bold on, height 222; align: horiz right; pattern: pattern solid, fore-color olive_ega', num_format_str="#,##0")
#         final_label_xf = ezxf('font: bold on, height 222; align: horiz left; pattern: pattern solid, fore-color light_green', num_format_str="#,##0")
#         final_xf = ezxf('font: bold on, height 222; align: horiz right; pattern: pattern solid, fore-color light_green', num_format_str="#,##0.00")
#     
#         ws0.row(0).height = 2500
#         # ws0.col(0).width = 6000
#         # ws0.col(1).width = 3000
#         # ws0.col(2).width = 3000
#         ws0.write_merge(0, 0, 0, 4, _("Results Summary for {0}".format(modelJSON['name'])), modelHeading_xf)
#         ws0.write_merge(1, 1, 0, 4, _("Overal System Statistics"), divider_1_xf)
#         ws0.write_merge(2, 2, 0, 4, _("Facility Statistics"), divider_2_xf)
#         ws0.write(3, 0, "Level", plain_label_xf)
#         ws0.write(3, 1, "Total", plain_label_xf)
#         ws0.write(3, 2, "Vaccinating", plain_label_xf)
#         row = 4
#         levels = modelJSON['orderedLevelList']
#         levelCount = modelJSON['fixedLocationCount']
#         vaccLevelCount = modelJSON['vaccinationLocationCount']
#         for level in levels:
#             if level in levelCount.keys():
#                 ws0.write(row, 0, level, plain_label_xf)
#                 ws0.write(row, 1, levelCount[level])
#                 if level in vaccLevelCount.keys():
#                     ws0.write(row, 2, vaccLevelCount[level])
#                 else:
#                     ws0.write(row, 2, 0)
#                 row += 1
#         ws0.write(row, 0, "All Levels", total_label_xf)
#         ws0.write(row, 1, levelCount['Total'], total_xf)
#         ws0.write(row, 2, vaccLevelCount['Total'], total_xf)
#         # ## Fit the column widths
#         # for i in range(4):
#         #    for j in range(row):
#                 
#        
#         
#         
# #         for level,count in m.getLevelList().items():
# #             ws0.write(row,1,level,plain_label_xf)
# #             ws0.write(row,2,count,plain_xf)
# #             row+=1
#         
#         # ws0.write_merge(1,1,0,4,_("Vaccine Results"),divider_1_xf)
#         
#         with uiSession.getLockedState() as state:
#             (fileKey, fullFileName) = \
#                 state.fs().makeNewFileInfo(shortName="%s.xls" % fname,
#                                            fileType='application/vnd.ms-excel',
#                                            deleteIfPresent=True)
#                 
#             wb.save(fullFileName)
#         return {'success':True, 'xlsfilename':fullFileName}
#     except Exception as e:
#         return {'success':False, 'msg':str(e)}


def summarizeTrucksByLevel(m, hr):
    truckFill = defaultdict(lambda : defaultdict(lambda : 0.0))
    truckCount = defaultdict(lambda : defaultdict(lambda : 0))

    for tr in hr.routesRpts.values():
        routeId = tr.RouteName
        route = m.routes[routeId]
        truck = route.TruckType
        # is level based on stop 0 or supplier?
        level = route.supplier().CATEGORY

        truckFill[level][truck] += tr.RouteFill
        truckCount[level][truck] += 1
    
    return {'fill' : truckFill, 'count' : truckCount}
        
        

def summarizeByLevel(hr):
    deliveryVol = defaultdict(lambda : 0)
    levelCount = defaultdict(lambda : 0)
    storageUtilizationSum = defaultdict(lambda : 0)
    storageUtilizationCount = defaultdict(lambda : 0)
    coolerVol = defaultdict(lambda : 0)
    usedCoolerVol = defaultdict(lambda : 0)

    for sr in hr.storesRpts.values():
        lev = sr.category
        deliveryVol[lev] += sr.tot_delivery_vol
        deliveryVol['Total'] += sr.tot_delivery_vol
        levelCount[lev] += 1
        levelCount['Total'] += 1

        if 'cooler' in sr.storage:
            cooler = sr.storage['cooler']
            if cooler.vol > 0.0:
                storageUtilizationSum[lev] += cooler.fillRatio * 100.0
                storageUtilizationCount[lev] += 1
                storageUtilizationSum['Total'] += cooler.fillRatio * 100.0
                storageUtilizationCount['Total'] += 1

                coolerVol[lev] += cooler.vol
                usedCoolerVol[lev] += cooler.vol_used
                coolerVol['Total'] += cooler.vol
                usedCoolerVol['Total'] += cooler.vol_used


    averagePeakUtilization = defaultdict(lambda : 0)
    for k,v in storageUtilizationSum.items():
        averagePeakUtilization[k] = v / storageUtilizationCount[k]

    ret = {}
    ret['deliveryVol'] = deliveryVol
    ret['levelCount'] = levelCount
    ret['averagePeakUtilization'] = averagePeakUtilization
    ret['coolerVol'] = coolerVol
    ret['usedCoolerVol'] = usedCoolerVol
    return ret

@bottle.route('/downloadXLS')
def downloadXLS(db, uiSession):
    shortname = _safeGetReqParam(bottle.request.params, 'shortname', isInt=False)
    xlsName = "%s.xlsx" % shortname
    with uiSession.getLockedState() as state:
        fullname = "%s/%s.xlsx" % (state.fs().workDir, shortname)
    return bottle.static_file(xlsName, os.path.dirname(fullname),
                              mimetype='application/vnd.ms-excel', download=xlsName)
    
@bottle.route('/json/downloadSVG',method='post')
def downloadSVG(db,uiSession):
    try:
        import json
        from xml.dom import minidom
        data = json.load(bottle.request.body)
        print '1243135315214234325124351341324554513431514513546514452465135315'
        print "Data = " + str(data)
        #print type(data.data)
        #svgXML = minidom.parse(str(data['data']))
        with uiSession.getLockedState() as state:
            #fullname = "{0}/test.svg".format(state.fs().workDir)
            fullname = "test.svg"
        with open(fullname,"wb") as f:
            f.write(data['data'])
        return {'success':True}
    
    except bottle.HTTPResponse:
        raise # bottle will handle this
    except Exception,e:
        result = {'success':False, 'type':'error', 'msg':_("Problem saving SVG XML: ") + str(e)}
        return result  
    
    
def getNumberOfLocationsFromModel(model):
    placeCount = 0
    vaccinatingCount = 0
    for storeId, store in model.stores.items():
        if store.isVaccinating():
            vaccinatingCount += 1
        if not store.isAttached():
            placeCount += 1
    return (placeCount, vaccinatingCount)

def getNumberOfLocationsByLevel(model):
    return model.getLevelCount()

class MicroCostDisplay():

    costGroups = {'solar': ('energy', 'total'), #'fuel/power',
                  'electric': ('energy', 'total'), #'fuel/power',
                  'diesel': ('energy', 'total'), #'fuel/power',
                  'gasoline': ('energy', 'total'), #'fuel/power',
                  'kerosene': ('energy', 'total'), #'fuel/power',
                  'petrol': ('energy', 'total'), #'fuel/power',
                  'ice': ('energy', 'total'), #'fuel/power',
                  'propane': ('energy', 'total'), #'fuel/power',
                  'StaffSalary': ('personnel', 'total'), # 'staff',
                  'PerDiem': ('t_perdiem', 't_total', 'total'), #'staff',
                  'TransitFareCost': ('t_fare', 't_total', 'total'), #'staff',
                  'Vaccines': ('vax', 'total'), #'cargo',
                  'BuildingCost': ('building', 'total'), #'equipment',
                  'FridgeAmort': ('s_amort', 's_total', 'total'), #'equipment',
                  'TruckAmort': ('t_amort', 't_total', 'total'), #'equipment',
                  'TruckMaint': ('t_maint', 't_total', 'total'), #'equipment',
                  'FridgeMaint': ('s_maint', 's_total', 'total'), #'equipment',
                  'SolarMaint': ('s_maint', 's_total', 'total'), #'equipment'
                  # other things I've found:
                  #'LaborCost':  ('personnel', 'total'),
                  #'Storage': ('s_maint', 's_total', 'total'),
                  #'Transport': ('t_maint', 't_total', 'total'),
                  }

    def microCostColumns(self):
        """
        set up our column headers and contents.
        this is a method so that the internationalization code is called every use.

        """
        if self.groups == 'all':
            columns = (
                (_("Energy"), ((_("Fuel and Electric"), 'energy'), )),
                (_("Storage"), ((_("Equipment Maintenance"), 's_maint'),
                                (_("Equipment Amortization"), 's_amort'),
                                (_("Total"), 's_total'))),
                (_("Transport"), ((_("Vehicle Maintenance"), 't_maint'),
                                  (_("Vehicle Amortization"), 't_amort'),
                                  (_("Fixed Fares"), 't_fare'),
                                  (_("Per Diems"), 't_perdiem'),
                                  (_("Total"), 't_total'))),
                (_("Personnel"), ((_("Total"), "personnel"),)),
                (_("Building"), ((_("Total"), "building"),)),
                (_("Vaccine Procurement"), ((_("Total"), "vax"),)),
                #        (_("Logistics"), # We don't have anything under this heading
                (_("Total Costs"), ((_("Including Procurement"), 'total'),)),
                )
        elif self.groups == 'stores':
            columns = (
                (_("Energy"), ((_("Total"), 'energy'), )),
                (_("Storage"), ((_("Equipment Maintenance"), 's_maint'),
                                (_("Equipment Amortization"), 's_amort'),
                                (_("Total"), 's_total'))),
                (_("Personnel"), ((_("Total"), "personnel"),)),
                (_("Building"), ((_("Total"), "building"),)),
                (_("Total Costs"), ((_("Including Procurement"), 'total'),)),
                )
        elif self.groups == 'routes':
            columns = (
                (_("Fuel"), ((_("Total"), 'energy'), )),
                (_("Transport"), ((_("Vehicle Maintenance"), 't_maint'),
                                  (_("Vehicle Amortization"), 't_amort'),
                                  (_("Fixed Fares"), 't_fare'),
                                  (_("Per Diems"), 't_perdiem'),
                                  (_("Total"), 't_total'))),
                #(_("Personnel"), ((_("Total"), "personnel"),)),
                (_("Total Costs"), ((_("Including Procurement"), 'total'),)),
                )
        else:
            raise(RuntimeError, "invalid costing display group request")

        return columns

    def __init__(self, groups='all'):
        """ what settings do I want here?
        allowed groups are 'all', 'stores', 'routes'
        """
        self.groups = groups
        self.columns = self.microCostColumns()
        self.columnCount = 0
        for h,c in self.columns:
            self.columnCount += len(c)
        
    def topHeaders(self, ss, style):
        for h,c in self.columns:
            ss.mergeCells(len(c))
            ss.set(h, style)
            ss.right(len(c))

    def detailHeaders(self, ss, style):
        for h, cList in self.columns:
            for cHead,data in cList:
                ss.postNext(cHead, style)

    def detailRow(self, ss, csr, style):
        info = defaultdict(lambda : 0)
        unknownCategories = []

        for ce in csr.costEntries:
            cc = ce.costCategory
            if cc.startswith('m1C_'):
                cc = cc[4:]
            else:
                # per joel if it doesn't start with m1C it doesn't count
                continue
            if cc not in self.costGroups:
                unknownCategories.append(cc)
                continue
            for grouping in self.costGroups[cc]:
                info[grouping] += ce.cost
        
        for h, cList in self.columns:
            for cHead, data in cList:
                ss.postNext(round(info[data], 2), style)
        ### STB While this is great for debugging, looks bad on the spreadsheet.
        #if len(unknownCategories) > 0:
        #    ss.set("unknown categories: %s"%unknownCategories)

    def levelSummary(self, costs, levels, ss):
        columnCount = 1 + self.columnCount
        ss.mergeCells(columnCount)
        ss.set(_("Costs By Level") + " (%s %s)"%(costs['year'], costs['currency']), 
               divider1Style)
        ss.nextRow()
        
        ss.rowHeight(30)
        ss.postNext("", labelStyleCenter)
        self.topHeaders(ss, labelStyleCenter)
        ss.nextRow()

        ss.rowHeight(30)
        ss.postNext("Level", labelStyle)
        self.detailHeaders(ss, labelStyle)
        ss.nextRow()

        for level in levels:
            if level not in costs['all']:
                continue
            ss.postNext(level, levelStyle)
            self.detailRow(ss, costs['all'][level], plainStyle)
            ss.nextRow()


def summarizeMicroCostLevels(costs, levels, ss):
    mcd = MicroCostDisplay()
    mcd.levelSummary(costs, levels, ss)

if 0:
    columnCount = 1 # include the row header (level)
    for h,c in columns:
        columnCount += len(c)

    ss.mergeCells(columnCount)
    ss.set(_("Costs By Level"), divider1Style)
    ss.nextRow()
    
    ss.rowHeight(30)
    ss.postNext("", labelStyleCenter)
    for h,c in columns:
        ss.mergeCells(len(c))
        ss.set(h, labelStyleCenter)
        ss.right(len(c))

    ss.nextRow()
    ss.rowHeight(30)
    ss.postNext("Level", labelStyle)
    for h, cList in columns:
        for cHead,data in cList:
            ss.postNext(cHead, labelStyle)
 
    ss.nextRow()

    levelCsrs = {}
    for csr in hr.costSummaryRecs:
        if csr.ReportingBranch in levels:
            levelCsrs[csr.ReportingBranch] = csr  # make it so I can find this later

    for level in levels:
        if level not in levelCsrs:
            continue
        levelInfo = defaultdict(lambda : 0)
        unknownCategories = []
        csr = levelCsrs[level]
        for ce in csr.costEntries:
            cc = ce.costCategory
            if cc.startswith('m1C_'):
                cc = cc[4:]
            if cc not in costGroups:
                unknownCategories.append(cc)
                continue
            for grouping in costGroups[cc]:
                levelInfo[grouping] += ce.cost
        
        ss.postNext(level, levelStyle)
        for h, cList in columns:
            for cHead, data in cList:
                ss.postNext(round(levelInfo[data], 2), plainStyle)
        if len(unknownCategories) > 0:
            ss.set("unknown categories: %s"%unknownCategories)
        ss.nextRow()

